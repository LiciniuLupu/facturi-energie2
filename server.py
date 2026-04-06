import os, base64, json, io, uuid
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__, static_folder=".")
CORS(app)
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
EXPORT_STORE = {}

DARK_BLUE="FF1F3864"; MED_BLUE="FF2E74B5"; LIGHT_BLUE="FFD6E4F0"
LIGHT_GRAY="FFF2F2F2"; YELLOW="FFFFF2CC"; GREEN_LIGHT="FFE2EFDA"
ORANGE_WARN="FFFCE4D6"; WHITE="FFFFFFFF"
MONTHS_RO=["Ianuarie","Februarie","Martie","Aprilie","Mai","Iunie","Iulie","August","Septembrie","Octombrie","Noiembrie","Decembrie"]

def tb():
    s=Side(style="thin",color="FFB8CCE4")
    return Border(left=s,right=s,top=s,bottom=s)

def cl(ws,row,col,val,bold=False,size=10,color="FF000000",bg=WHITE,halign="left",indent=0):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(name="Arial",size=size,bold=bold,color=color)
    c.fill=PatternFill("solid",fgColor=bg)
    c.alignment=Alignment(horizontal=halign,vertical="center",indent=indent)
    c.border=tb()
    return c

def make_excel(invoices, year, title, source_note):
    wb=Workbook(); ws=wb.active; ws.title="Consum Lunar"
    ws.column_dimensions["A"].width=45
    ws.column_dimensions["B"].width=14
    ws.column_dimensions["C"].width=8
    ws.column_dimensions["D"].width=16
    ws.column_dimensions["E"].width=28

    logo_path=os.path.join(os.path.dirname(os.path.abspath(__file__)),"logo_small.png")
    if os.path.exists(logo_path):
        try:
            img=XLImage(logo_path); img.anchor="F1"; ws.add_image(img)
            ws.row_dimensions[1].height=55
        except Exception: pass

    ws.merge_cells("A1:E1")
    c=ws["A1"]; c.value=title
    c.font=Font(name="Arial",size=14,bold=True,color="FFFFFFFF")
    c.fill=PatternFill("solid",fgColor=DARK_BLUE)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)

    ws.merge_cells("A2:E2")
    c=ws["A2"]; c.value=source_note
    c.font=Font(name="Arial",size=10,color="FF595959")
    c.fill=PatternFill("solid",fgColor=LIGHT_GRAY)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[2].height=16
    ws.row_dimensions[3].height=6

    for col,h in enumerate(["Nr.","Luna","An","Consum (kWh)","Obs."],1):
        cl(ws,4,col,h,bold=True,size=11,color="FFFFFFFF",bg=MED_BLUE,halign="center")
    ws.row_dimensions[4].height=18

    # Group by year+month, sorted
    by_ym = {}
    for inv in invoices:
        m=int(inv.get("luna",0)); an=int(inv.get("an",0))
        if 1<=m<=12 and an>=2020:
            by_ym[(an,m)]={"kwh":float(inv.get("kwh",0)),"obs":inv.get("obs","") or ""}

    data_rows = sorted(by_ym.items())  # [(year,month), data]

    # For solar calc: find dominant year
    years_present = sorted(set(an for (an,m) in by_ym.keys()))
    dominant_year = year  # use detected year

    # Months present for dominant year
    months_in_year = {m: by_ym[(dominant_year,m)] for (an,m) in by_ym if an==dominant_year}
    complete_months = [m for m,v in months_in_year.items() if v["kwh"]>0 and "par" not in v["obs"].lower()]
    missing_months = [MONTHS_RO[m-1] for m in range(1,13) if m not in months_in_year]

    FDR=5
    for i,((an,m),d) in enumerate(data_rows):
        row=FDR+i; bg=LIGHT_BLUE if i%2==0 else WHITE
        for col,v in enumerate([i+1,MONTHS_RO[m-1],an,d["kwh"],d["obs"]],1):
            cl(ws,row,col,v,bold=(col==4),bg=bg,halign="center" if col<5 else "left")
        ws.row_dimensions[row].height=16

    TR=FDR+len(data_rows)
    ws.merge_cells(f"A{TR}:C{TR}")
    c=ws[f"A{TR}"]; c.value="TOTAL"
    c.font=Font(name="Arial",size=11,bold=True,color="FFFFFFFF")
    c.fill=PatternFill("solid",fgColor=DARK_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=tb()
    c=ws[f"D{TR}"]; c.value=f"=SUM(D{FDR}:D{TR-1})" if data_rows else 0
    c.font=Font(name="Arial",size=11,bold=True,color="FFFFFFFF")
    c.fill=PatternFill("solid",fgColor=DARK_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=tb()
    c=ws[f"E{TR}"]; c.fill=PatternFill("solid",fgColor=DARK_BLUE); c.border=tb()
    ws.row_dimensions[TR].height=18

    SP=TR+1; ws.row_dimensions[SP].height=10
    SHR=SP+1
    ws.merge_cells(f"A{SHR}:E{SHR}")
    c=ws[f"A{SHR}"]; c.value=f"ESTIMARE NECESAR PARC FOTOVOLTAIC — {dominant_year}"
    c.font=Font(name="Arial",size=12,bold=True,color="FFFFFFFF")
    c.fill=PatternFill("solid",fgColor=MED_BLUE)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[SHR].height=18

    NR=SHR+1
    if complete_months:
        names=[MONTHS_RO[m-1] for m in sorted(complete_months)]
        basis=f"Calcul bazat pe lunile complete din {dominant_year}: {', '.join(names)}"
    else:
        basis=f"Calcul bazat pe toate lunile disponibile din {dominant_year}."
    ws.merge_cells(f"A{NR}:E{NR}")
    c=ws[f"A{NR}"]; c.value=basis
    c.font=Font(name="Arial",size=9,color="FF595959",italic=True)
    c.fill=PatternFill("solid",fgColor=LIGHT_GRAY)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[NR].height=14

    PHR=NR+1
    for col,h in enumerate(["Parametru","Valoare","Unitate","Explicatie"],1):
        cl(ws,PHR,col,h,bold=True,size=11,color="FFFFFFFF",bg=MED_BLUE,halign="center")
    ws.row_dimensions[PHR].height=18

    r=PHR+1; AVG_R=r

    # Find row indices for complete months in Excel
    complete_rows = [FDR+i for i,((an,m),_) in enumerate(data_rows) if an==dominant_year and m in complete_months]
    if complete_rows:
        avg_val=f"=AVERAGE({','.join([f'D{rr}' for rr in complete_rows])})"
    elif data_rows:
        avg_val=f"=AVERAGE(D{FDR}:D{TR-1})"
    else:
        avg_val=0

    solar_params=[
        ("Consum mediu lunar (luni complete)",avg_val,"kWh/luna","Media lunilor complete disponibile",YELLOW),
        ("Consum anual estimat",f"=ROUND(B{r}*12,0)","kWh/an","Consum mediu lunar x 12 luni",YELLOW),
        ("Ore de varf solare (Romania)",3.8,"h/zi","Medie nationala (~3.5-4.2 h/zi)",LIGHT_GRAY),
        ("Productie specifica anuala (per kWp)",f"=ROUND(B{r+2}*365,0)","kWh/kWp/an","Ore varf/zi x 365 zile",WHITE),
        ("Factor utilizare sistem",0.8,"—","Eficienta globala sistem ~80%",LIGHT_GRAY),
        ("Productie neta per kWp instalat",f"=ROUND(B{r+3}*B{r+4},0)","kWh net/kWp/an","Productie specifica x factor utilizare",WHITE),
    ]
    for label,val,unit,expl,bg in solar_params:
        for col,v in enumerate([label,val,unit,expl],1):
            cl(ws,r,col,v,bg=bg,halign="center" if col in (2,3) else "left",indent=1 if col==1 else 0)
        ws.row_dimensions[r].height=16; r+=1

    KWP_R=r; NET_R=r-1; ANN_R=AVG_R+1
    for col,v in enumerate(["NECESAR PUTERE INSTALATA (kWp)",f"=ROUND(B{ANN_R}/B{NET_R},1)","kWp","Consum anual / productie neta per kWp"],1):
        cl(ws,KWP_R,col,v,bold=True,bg=GREEN_LIGHT,halign="center" if col==2 else "left",indent=1 if col==1 else 0)
    ws.row_dimensions[KWP_R].height=20; r+=1

    for label,val,unit,expl,bg in [
        ("Nr. panouri estimat (400 Wp/panou)",f"=CEILING(B{KWP_R}*1000/400,1)","panouri","Putere instalata kWp / 0.4 kWp/panou",WHITE),
        ("Suprafata necesara estimata",f"=ROUND(B{r}*2,1)","m2","~2 m2 per panou",LIGHT_GRAY),
    ]:
        for col,v in enumerate([label,val,unit,expl],1):
            cl(ws,r,col,v,bg=bg,halign="center" if col in (2,3) else "left",indent=1 if col==1 else 0)
        ws.row_dimensions[r].height=16; r+=1

    r+=1
    if missing_months:
        ws.merge_cells(f"A{r}:E{r}")
        c=ws[f"A{r}"]; c.value=f"ATENTIE - LUNI LIPSA DIN ARHIVA ({dominant_year})"
        c.font=Font(name="Arial",size=11,bold=True,color="FF833C00")
        c.fill=PatternFill("solid",fgColor=ORANGE_WARN)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[r].height=18; r+=1
        for line in [f"Luni lipsa pentru {dominant_year}: {', '.join(missing_months)}.",
                     "Estimarea fotovoltaica se bazeaza pe lunile disponibile.",
                     "Pentru un calcul exact, incarcati si facturile lipsa."]:
            ws.merge_cells(f"A{r}:E{r}")
            c=ws[f"A{r}"]; c.value=line
            c.font=Font(name="Arial",size=10,color="FF833C00")
            c.fill=PatternFill("solid",fgColor=ORANGE_WARN)
            c.alignment=Alignment(horizontal="left",vertical="center",indent=2); ws.row_dimensions[r].height=15; r+=1
    else:
        ws.merge_cells(f"A{r}:E{r}")
        c=ws[f"A{r}"]; c.value=f"Toate cele 12 luni din {dominant_year} sunt prezente. Estimare bazata pe consum complet anual."
        c.font=Font(name="Arial",size=10,bold=True,color="FF375623")
        c.fill=PatternFill("solid",fgColor=GREEN_LIGHT)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[r].height=16

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


@app.route("/")
def index(): return send_from_directory(".","index.html")

@app.route("/logo_small.png")
def logo(): return send_from_directory(".","logo_small.png")

@app.route("/analyze", methods=["POST"])
def analyze():
    if not ANTHROPIC_API_KEY:
        return jsonify({"error":"ANTHROPIC_API_KEY nu este setat."}),500
    files=request.files.getlist("files")
    if not files: return jsonify({"error":"Nu au fost trimise fisiere."}),400

    client=anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    parts=[{"type":"text","text":"""Esti expert in analiza facturi energie electrica Romania.
Analizeaza TOATE documentele atasate. Pentru fiecare factura extrage:
1. Luna si anul PERIOADEI DE CONSUM (nu data emiterii facturii!)
2. Consumul in kWh (cauta: consum activ, energie activa, diferenta indecsi)
3. Daca luna e partiala (<28 zile de consum), noteaza in campul obs

IMPORTANT: Nu presupune anul - citeste-l exact din fiecare factura.
Pot exista facturi din ani diferiti - include-le pe toate.

Raspunde EXCLUSIV cu JSON valid, fara markdown, fara backticks, fara alt text:
{"invoices":[{"luna":1,"an":2024,"kwh":5200,"obs":""},{"luna":2,"an":2024,"kwh":4800,"obs":"Luna partiala"}]}"""}]

    for f in files:
        raw=f.read(); b64=base64.standard_b64encode(raw).decode(); mime=f.content_type or "application/octet-stream"
        if mime.startswith("image/"): parts.append({"type":"image","source":{"type":"base64","media_type":mime,"data":b64}})
        elif mime=="application/pdf" or f.filename.lower().endswith(".pdf"): parts.append({"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}})
        else:
            try: parts.append({"type":"text","text":f"\n---{f.filename}---\n{raw.decode('utf-8',errors='replace')}"})
            except: pass
    try:
        msg=client.messages.create(model="claude-sonnet-4-5",max_tokens=2048,messages=[{"role":"user","content":parts}])
        txt="".join(b.text for b in msg.content if hasattr(b,"text")).strip()
        start=txt.find("{"); end=txt.rfind("}")+1
        if start>=0 and end>start: txt=txt[start:end]
        parsed=json.loads(txt)
        invs=[i for i in (parsed.get("invoices") or []) if int(i.get("an",0))>=2020]

        # Detect dominant year: year with most invoices
        from collections import Counter
        year_counts=Counter(int(i["an"]) for i in invs)
        dominant_year=year_counts.most_common(1)[0][0] if year_counts else 2024

        token=str(uuid.uuid4())
        EXPORT_STORE[token]={
            "invoices":invs,"year":dominant_year,
            "title":request.form.get("title","CONSUM ENERGIE ELECTRICA"),
            "source_note":request.form.get("source_note","Sursa: Facturi energie electrica")
        }
        return jsonify({"invoices":invs,"year":dominant_year,"token":token})
    except json.JSONDecodeError as e: return jsonify({"error":f"Raspuns invalid AI: {e}"}),500
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/export/<token>", methods=["GET"])
def export(token):
    data=EXPORT_STORE.get(token)
    if not data:
        return "Token invalid sau expirat. Reia analiza.", 404
    yr=int(data.get("year",2024))
    xlsx=make_excel(data.get("invoices",[]),yr,data.get("title","CONSUM ENERGIE ELECTRICA"),data.get("source_note","Sursa: Facturi energie electrica"))
    return send_file(io.BytesIO(xlsx),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Consum_Energie_{yr}.xlsx")

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    print(f"Server pornit pe http://localhost:{port}")
    app.run(host="0.0.0.0",port=port,debug=False)
